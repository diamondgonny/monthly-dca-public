from gspread import Worksheet
import datetime
import json
import os
import openpyxl
import time


def get_exchange_rate_from_google_sheet(worksheet: Worksheet):
    # 간혹 float 타입의 환율 대신, 문자열인 '로딩 중...'을 입력받아 생기는 문제를 보완하고자 backup
    exchange_rate: float = 0
    while True:
        try:
            exchange_rate = float(worksheet.cell(2, 21).value.lstrip('₩').replace(',', ''))
        except ValueError:
            print("환율을 '로드 중'일 때 불러왔기 떄문에 ValueError 발생... 5초 후 다시 시도하겠습니다.")
            time.sleep(5)
        if exchange_rate:
            break

    return exchange_rate


def get_pointer_ticker_plate_from_google_sheet(worksheet1: Worksheet):
    # row_base_pointer, row_stack_pointer의 기준은 선 안의 색칠된 칸
    row_stack_pointer_last = len(worksheet1.get_all_values())  # end of rows
    nested_raw_ticker_list = worksheet1.batch_get([f'A1:A{row_stack_pointer_last}'])
    flat_raw_ticker_list = [element[0] if element else '' for sublist in nested_raw_ticker_list for element in sublist]
    """flat_raw_ticker_list = ['[1] ', 'TLT', 'BRK.B', 'XLU', 'AMZN', 'GOOG', 'MSFT', 'Sum', '', '[2] ', 'VTI', 'VEA', 'VWO', 'DBC', 'IAU', 'EDV', 'LTPZ', 'VCLT', 'EMLC', 'Sum', '', '', '[3] ', 'GLDM', 'PDBC', 'IEMG', 'VT', 'ITOT', 'MCHI', 'VGK', 'EZA', 'EWZ', 'EWW', 'EWJ', 'EWH', 'EWC', 'EMB', 'TIP', 'HYG', 'LEMB', 'VGLT', 'Sum']"""

    row_base_pointer_list = [(i + 1) + 1 for i, ticker in enumerate(flat_raw_ticker_list) if '[' in ticker and ']' in ticker]
    row_stack_pointer_list = [i + 1 for i, ticker in enumerate(flat_raw_ticker_list) if 'Sum' in ticker]
    portfolio_ticker_list = [flat_raw_ticker_list[rbp - 1:rsp - 1] for rbp, rsp in zip(row_base_pointer_list, row_stack_pointer_list)]
    print(portfolio_ticker_list)
    """portfolio_ticker_list = [['TLT', 'BRK.B', 'XLU', 'AMZN', 'GOOG', 'MSFT'], ['VTI', 'VEA', 'VWO', 'DBC', 'IAU', 'EDV', 'LTPZ', 'VCLT', 'EMLC'], ['GLDM', 'PDBC', 'IEMG', 'VT', 'ITOT', 'MCHI', 'VGK', 'EZA', 'EWZ', 'EWW', 'EWJ', 'EWH', 'EWC', 'EMB', 'TIP', 'HYG', 'LEMB', 'VGLT']]"""

    row_pointer_dict = {}
    for i, _ in enumerate(row_base_pointer_list):
        row_pointer_dict.update({f'row_base_pointer_{i + 1}': row_base_pointer_list[i]})
        row_pointer_dict.update({f'row_stack_pointer_{i + 1}': row_stack_pointer_list[i]})
    """row_pointer_dict = {'row_base_pointer_1': 2, 'row_stack_pointer_1': 8, 'row_base_pointer_2': 11, 'row_stack_pointer_2': 20, 'row_base_pointer_3': 24, 'row_stack_pointer_3': 42}"""
    plate_portfolio_mkt_qty_dict = {}
    for i, _ in enumerate(row_base_pointer_list):
        plate_portfolio_mkt_qty_dict.update({f'portfolio_mkt_qty_{i + 1}': {ticker: ['', 0] for ticker in portfolio_ticker_list[i]}})
    """plate_portfolio_mkt_qty_dict = {'portfolio_mkt_qty_1': {'TLT': ['', 0], 'BRK.B': ['', 0], 'XLU': ['', 0], 'AMZN': ['', 0], 'GOOG': ['', 0], 'MSFT': ['', 0]}, 'portfolio_mkt_qty_2': {'VTI': ['', 0], 'VEA': ['', 0], 'VWO': ['', 0], 'DBC': ['', 0], 'IAU': ['', 0], 'EDV': ['', 0], 'LTPZ': ['', 0], 'VCLT': ['', 0], 'EMLC': ['', 0]}, 'portfolio_mkt_qty_3': {'GLDM': ['', 0], 'PDBC': ['', 0], 'IEMG': ['', 0], 'VT': ['', 0], 'ITOT': ['', 0], 'MCHI': ['', 0], 'VGK': ['', 0], 'EZA': ['', 0], 'EWZ': ['', 0], 'EWW': ['', 0], 'EWJ': ['', 0], 'EWH': ['', 0], 'EWC': ['', 0], 'EMB': ['', 0], 'TIP': ['', 0], 'HYG': ['', 0], 'LEMB': ['', 0], 'VGLT': ['', 0]}}"""

    return row_pointer_dict, portfolio_ticker_list, plate_portfolio_mkt_qty_dict


def send_material_to_plate_and_do_ticker_missing_check(material_mkt_qty_dict: dict, plate_portfolio_mkt_qty_dict: dict):
    # material_mkt_qty_dict -> plate_portfolio_mkt_qty_dict, 전산망에선 조회되나(현재보유중) Google Sheet 기록에서 누락종목이 있는지 검사해주는 함수
    for i, _ in enumerate(plate_portfolio_mkt_qty_dict):
        for ticker in plate_portfolio_mkt_qty_dict[f'portfolio_mkt_qty_{i + 1}'].keys():
            if ticker in material_mkt_qty_dict:
                plate_portfolio_mkt_qty_dict[f'portfolio_mkt_qty_{i + 1}'][ticker][0] = material_mkt_qty_dict[ticker][0]
                plate_portfolio_mkt_qty_dict[f'portfolio_mkt_qty_{i + 1}'][ticker][1] = material_mkt_qty_dict[ticker][1]
                del material_mkt_qty_dict[ticker]
    print('Verified Possession:\n', plate_portfolio_mkt_qty_dict)
    print('Missing Possession (which Google Sheet missed out):\n', material_mkt_qty_dict)

    if not material_mkt_qty_dict:
        print('[Pass] Checked for missing ticker by comparing from account.\n')
    else:
        raise Exception(f'{list(material_mkt_qty_dict.keys())} 보유기록 누락 (Google Sheet에 기록되지 않은 종목, 업데이트 요망)')

    return plate_portfolio_mkt_qty_dict


def do_ticker_duplicate_check(portfolio_ticker_list: list):
    # Google Sheet 기록에서 중복된 종목이 있는지 검사해주는 함수
    flat_ticker_list = [ticker for portfolio_n_ticker_list in portfolio_ticker_list for ticker in portfolio_n_ticker_list]
    """flat_ticker_list = ['TLT', 'BRK.B', 'XLU', 'AMZN', 'GOOG', 'MSFT', 'VTI', 'VEA', 'VWO', 'DBC', 'IAU', 'EDV', 'LTPZ', 'VCLT', 'EMLC', 'GLDM', 'PDBC', 'IEMG', 'VT', 'ITOT', 'MCHI', 'VGK', 'EZA', 'EWZ', 'EWW', 'EWJ', 'EWH', 'EWC', 'EMB', 'TIP', 'HYG', 'LEMB', 'VGLT']"""
    duplicates = set(ticker for ticker in flat_ticker_list if flat_ticker_list.count(ticker) > 1)  # duplicates = {'(ticker)', ...}

    if not duplicates:
        print('[Pass] Checked for ticker duplicate check in portfolio_mkt_qty_dict.\n')
    else:
        raise Exception(f'{duplicates} 보유종목 중복 (Google Sheet 업데이트 요망)')


def get_mkt_info_from_ticker_book(portfolio_mkt_qty_dict: dict, pf_code: int):
    # 전산망에서 조회되지 않은 종목(현재보유수량이 0인 종목)에 market 데이터를 보충해주는 함수
    print('Loading...')
    xlsx_file_name = os.getcwd() + '/overseas_stock_code/overseas_stock_code(all).xlsx'  # 엑셀파일 열기, 엑셀 파일의 첫번째 시트 추출하기
    book_xl = openpyxl.load_workbook(xlsx_file_name)
    worksheet_xl = book_xl.worksheets[0]

    portfolio_mkt_qty_pf_code = portfolio_mkt_qty_dict[f'portfolio_mkt_qty_{pf_code}']
    for ticker, mkt_qty_list in portfolio_mkt_qty_pf_code.items():
        if mkt_qty_list[0] == '':
            for cell in worksheet_xl['E']:
                if ticker == cell.value:
                    market: str = worksheet_xl.cell(row=cell.row, column=cell.column - 2).value
                    market = {'AMS': 'AMEX', 'NAS': 'NASD', 'NYS': 'NYSE'}.get(market, market)
                    mkt_qty_list[0] = market
                    break
    print(f"Received market data from Ticker Book to 'portfolio_mkt_qty_{pf_code}'.\n")


def print_portfolio_mkt_qty_from_google_sheet(worksheet: Worksheet, row_pointer_dict: dict, portfolio_mkt_qty_dict: dict):
    for i, _ in enumerate(portfolio_mkt_qty_dict):
        print(worksheet.cell(row_pointer_dict[f'row_base_pointer_{i + 1}'] - 1, 1).value)
        print('row_base_pointer : {0}'.format(row_pointer_dict[f'row_base_pointer_{i + 1}']))
        print('row_stack_pointer : {0}'.format(row_pointer_dict[f'row_stack_pointer_{i + 1}']))
        print(portfolio_mkt_qty_dict[f'portfolio_mkt_qty_{i + 1}'])
        print()


def backup_portfolio_mkt_qty_info_to_json_file(row_pointer_dict: dict, exchange_rate: float, buyable_cash_dollar_sum: float,
                                               buyable_cash_dollar_list: list, portfolio_mkt_qty_dict: dict):
    update_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    portfolio_mkt_qty_info = {
        "update_date": update_date,
        "row_pointer_dict": row_pointer_dict,
        "exchange_rate": exchange_rate,
        "buyable_cash_dollar_sum": buyable_cash_dollar_sum,
        "buyable_cash_dollar_list": buyable_cash_dollar_list,
        "portfolio_mkt_qty_dict": portfolio_mkt_qty_dict
    }
    with open('./config/backup-portfolio-mkt-qty-info.json', 'w', encoding='UTF-8') as f:
        json.dump(portfolio_mkt_qty_info, f, indent=2, sort_keys=False)
        print('Saved the portfolio backup data in ./config/backup-portfolio-mkt-qty-info.json.\n')


def put_mkt_and_qty_to_google_sheet(worksheet: Worksheet, row_pointer_dict: dict, portfolio_mkt_qty_dict: dict, pf_mkt_req_list: list):
    ranges = []
    values = []
    # mkt (selected portfolio : pf_mkt_req_list)
    for mkt_val in pf_mkt_req_list:
        ranges.append(f"D{row_pointer_dict[f'row_base_pointer_{mkt_val}']}:D{row_pointer_dict[f'row_stack_pointer_{mkt_val}'] - 1}")
        values.append([[mkt_qty_list[0]] for ticker, mkt_qty_list in portfolio_mkt_qty_dict[f'portfolio_mkt_qty_{mkt_val}'].items()])
    # qty (all portfolios)
    for qty_i, _ in enumerate(portfolio_mkt_qty_dict):
        ranges.append(f"G{row_pointer_dict[f'row_base_pointer_{qty_i + 1}']}:G{row_pointer_dict[f'row_stack_pointer_{qty_i + 1}'] - 1}")
        values.append([[mkt_qty_list[1]] for ticker, mkt_qty_list in portfolio_mkt_qty_dict[f'portfolio_mkt_qty_{qty_i + 1}'].items()])

    update_requests = [{'range': r, 'values': v} for r, v in zip(ranges, values)]
    worksheet.batch_update(update_requests)
    print('Updated market, current qty data(via portfolio_mkt_qty) from account to Google Sheet.')


def put_cash_to_google_sheet(worksheet: Worksheet, row_pointer_dict: dict, exchange_rate: float, expected_final_balance_dollar: float,
                             buyable_cash_dollar_list: list):
    for i, _ in enumerate(buyable_cash_dollar_list):
        worksheet.update_cell(row_pointer_dict[f'row_stack_pointer_{i + 1}'], 16, int(buyable_cash_dollar_list[i] * exchange_rate))
    worksheet.update_cell(2, 20, expected_final_balance_dollar * exchange_rate)
    print('Updated buyable_cash_won for portfolios, expected_final_balance_won from account to Google Sheet.\n')
